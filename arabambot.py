import requests
from bs4 import BeautifulSoup
from  openpyxl import *
import time
kitap = Workbook()
print("""
****************************************************

            ARABAM.COM BOT UYGULAMASI
                V.1.0.0
                 FTHSN

 
****************************************************
""")
arac = input("Hangi aracın verilerini almak istersiniz:")
kac = int(input("Kac sayfa veri çekmek istersiniz(Her sayfada 20 ilan):"))
i=1
while i <=kac:
    marka = arac+"?page"+str(i)
    url = "https://www.arabam.com/ikinci-el/otomobil/"+marka
    yeniurl = requests.get(url)
    soup = BeautifulSoup(yeniurl.content,"lxml")
    ne = soup.find_all("tr",attrs={"class":"listing-list-item pr should-hover bg-white"})
    say = 0
    i=i+1
    for a in ne:
        model = a.find("h3",attrs={"class":"crop-after"}).text
        baslik =a.find("td",attrs={"class":"horizontal-half-padder-minus pr"}).text
        Yıl = a.find_all("td",attrs={"class":"listing-text pl8 pr8 tac pr"})[0].text
        Km = a.find_all("td",attrs={"class":"listing-text pl8 pr8 tac pr"})[1].text
        Renk = a.find_all("td",attrs={"class":"listing-text pl8 pr8 tac pr"})[2].text
        Fiyat = a.find("td",attrs={"class":"pl8 pr8 tac pr"}).text
        say +=1
        sheet = kitap.active
        sheet.append((model,baslik,Yıl,Km,Renk,Fiyat))
        kitap.save("{}.xlsx".format(arac))
        kitap.close
wb = load_workbook("{}.xlsx".format(arac))
ws = wb.active

kactane = len(ws['A'])
print("""Toplamda {} ilan {}.xlsx excel dosyasına kaydedildi.\nProgramdan çıkılıyor...""".format(kactane,arac))
time.sleep(5)
